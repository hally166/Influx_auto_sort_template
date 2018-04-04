""" 
Latin Square Generator for BD Influx running Sortware v1.2.0.142
v1.81 APR 2018
Python 3.6 (2017)
Author : Christopher Hall, Wellcome Sanger Institute, ch15@sanger.ac.uk
Latin square sort designs by Stephan Lorenz, Wellcome Sanger Institute 
License : GPLv3 https://www.gnu.org/licenses/gpl-3.0.html
"""

#import dependencies
import glob
import os
import re
import fileinput
import sys
import time
import shutil
#dictionaries of the latin square schemes
two_pop_96w_dict = {"A1":"gate1", "A2":"gate2", "A3":"gate1", "A4":"gate2", "A5":"gate1", "A6":"gate2", "A7":"gate1", "A8":"gate2", "A9":"gate1", "A10":"gate2", "A11":"gate1", "A12":"gate2", "B1":"gate1", "B2":"gate2", "B3":"gate1", "B4":"gate2", "B5":"gate1", "B6":"gate2", "B7":"gate1", "B8":"gate2", "B9":"gate1", "B10":"gate2", "B11":"gate1", "B12":"gate2", "C1":"gate2", "C2":"gate1", "C3":"gate2", "C4":"gate1", "C5":"gate2", "C6":"gate1", "C7":"gate2", "C8":"gate1", "C9":"gate2", "C10":"gate1", "C11":"gate2", "C12":"gate1", "D1":"gate2", "D2":"gate1", "D3":"gate2", "D4":"gate1", "D5":"gate2", "D6":"gate1", "D7":"gate2", "D8":"gate1", "D9":"gate2", "D10":"gate1", "D11":"gate2", "D12":"gate1", "E1":"gate1", "E2":"gate2", "E3":"gate1", "E4":"gate2", "E5":"gate1", "E6":"gate2", "E7":"gate1", "E8":"gate2", "E9":"gate1", "E10":"gate2", "E11":"gate1", "E12":"gate2", "F1":"gate1", "F2":"gate2", "F3":"gate1", "F4":"gate2", "F5":"gate1", "F6":"gate2", "F7":"gate1", "F8":"gate2", "F9":"gate1", "F10":"gate2", "F11":"gate1", "F12":"gate2", "G1":"gate2", "G2":"gate1", "G3":"gate2", "G4":"gate1", "G5":"gate2", "G6":"gate1", "G7":"gate2", "G8":"gate1", "G9":"gate2", "G10":"gate1", "G11":"gate2", "G12":"gate1", "H1":"gate2", "H2":"gate1", "H3":"gate2", "H4":"gate1", "H5":"gate2", "H6":"gate1", "H7":"gate2", "H8":"gate1", "H9":"", "H10":"", "H11":"", "H12":""}
three_pop_96w_dict = {"A1":"gate1", "A2":"gate1", "A3":"gate2", "A4":"gate2", "A5":"gate3", "A6":"gate3", "A7":"gate2", "A8":"gate2", "A9":"gate3", "A10":"gate3", "A11":"gate1", "A12":"gate1", "B1":"gate2", "B2":"gate3", "B3":"gate1", "B4":"gate3", "B5":"gate1", "B6":"gate2", "B7":"gate1", "B8":"gate3", "B9":"gate1", "B10":"gate2", "B11":"gate2", "B12":"gate3", "C1":"gate3", "C2":"gate2", "C3":"gate3", "C4":"gate1", "C5":"gate2", "C6":"gate1", "C7":"gate3", "C8":"gate1", "C9":"gate2", "C10":"gate1", "C11":"gate3", "C12":"gate2", "D1":"gate1", "D2":"gate2", "D3":"gate2", "D4":"gate3", "D5":"gate3", "D6":"gate1", "D7":"gate2", "D8":"gate3", "D9":"gate3", "D10":"gate1", "D11":"gate1", "D12":"gate2", "E1":"gate3", "E2":"gate1", "E3":"gate3", "E4":"gate1", "E5":"gate2", "E6":"gate2", "E7":"gate3", "E8":"gate1", "E9":"gate2", "E10":"gate2", "E11":"gate3", "E12":"gate1", "F1":"gate2", "F2":"gate3", "F3":"gate1", "F4":"gate2", "F5":"gate1", "F6":"gate3", "F7":"gate1", "F8":"gate2", "F9":"gate1", "F10":"gate3", "F11":"gate2", "F12":"gate3", "G1":"gate3", "G2":"gate2", "G3":"gate3", "G4":"gate1", "G5":"gate2", "G6":"gate1", "G7":"gate3", "G8":"gate1", "G9":"gate2", "G10":"gate1", "G11":"gate3", "G12":"gate2", "H1":"gate1", "H2":"gate2", "H3":"gate2", "H4":"gate3", "H5":"gate3", "H6":"gate1", "H7":"gate2", "H8":"gate3", "H9":"", "H10":"", "H11":"", "H12":""}
four_pop_96w_dict = {"A1":"gate1", "A2":"gate2", "A3":"gate2", "A4":"gate2", "A5":"gate3", "A6":"gate3", "A7":"gate4", "A8":"gate1", "A9":"gate1", "A10":"gate2", "A11":"gate2", "A12":"gate3", "B1":"gate3", "B2":"gate4", "B3":"gate1", "B4":"gate3", "B5":"gate1", "B6":"gate4", "B7":"gate3", "B8":"gate2", "B9":"gate4", "B10":"gate1", "B11":"gate3", "B12":"gate1", "C1":"gate3", "C2":"gate2", "C3":"gate4", "C4":"gate1", "C5":"gate4", "C6":"gate2", "C7":"gate1", "C8":"gate3", "C9":"gate2", "C10":"gate4", "C11":"gate1", "C12":"gate4", "D1":"gate4", "D2":"gate3", "D3":"gate3", "D4":"gate4", "D5":"gate2", "D6":"gate1", "D7":"gate2", "D8":"gate1", "D9":"gate3", "D10":"gate2", "D11":"gate4", "D12":"gate2", "E1":"gate2", "E2":"gate2", "E3":"gate4", "E4":"gate3", "E5":"gate4", "E6":"gate4", "E7":"gate1", "E8":"gate2", "E9":"gate2", "E10":"gate3", "E11":"gate3", "E12":"gate4", "F1":"gate1", "F2":"gate3", "F3":"gate1", "F4":"gate4", "F5":"gate3", "F6":"gate2", "F7":"gate4", "F8":"gate1", "F9":"gate3", "F10":"gate1", "F11":"gate4", "F12":"gate3", "G1":"gate4", "G2":"gate1", "G3":"gate4", "G4":"gate2", "G5":"gate1", "G6":"gate3", "G7":"gate2", "G8":"gate4", "G9":"gate1", "G10":"gate4", "G11":"gate2", "G12":"gate1", "H1":"gate3", "H2":"gate4", "H3":"gate2", "H4":"gate1", "H5":"gate2", "H6":"gate1", "H7":"gate3", "H8":"gate3", "H9":"", "H10":"", "H11":"", "H12":""}
five_pop_96w_dict = {"A1":"gate1", "A2":"gate2", "A3":"gate3", "A4":"gate4", "A5":"gate5", "A6":"gate4", "A7":"gate1", "A8":"gate2", "A9":"gate3", "A10":"gate4", "A11":"gate5", "A12":"gate2", "B1":"gate4", "B2":"gate5", "B3":"gate2", "B4":"gate1", "B5":"gate3", "B6":"gate1", "B7":"gate4", "B8":"gate5", "B9":"gate2", "B10":"gate1", "B11":"gate3", "B12":"gate5", "C1":"gate5", "C2":"gate3", "C3":"gate1", "C4":"gate2", "C5":"gate4", "C6":"gate2", "C7":"gate5", "C8":"gate3", "C9":"gate1", "C10":"gate2", "C11":"gate4", "C12":"gate3", "D1":"gate2", "D2":"gate4", "D3":"gate5", "D4":"gate3", "D5":"gate1", "D6":"gate3", "D7":"gate2", "D8":"gate4", "D9":"gate5", "D10":"gate3", "D11":"gate1", "D12":"gate4", "E1":"gate3", "E2":"gate1", "E3":"gate4", "E4":"gate5", "E5":"gate2", "E6":"gate5", "E7":"gate3", "E8":"gate1", "E9":"gate4", "E10":"gate5", "E11":"gate2", "E12":"gate1", "F1":"gate4", "F2":"gate5", "F3":"gate2", "F4":"gate1", "F5":"gate3", "F6":"gate1", "F7":"gate4", "F8":"gate5", "F9":"gate2", "F10":"gate1", "F11":"gate3", "F12":"gate5", "G1":"gate5", "G2":"gate3", "G3":"gate1", "G4":"gate2", "G5":"gate4", "G6":"gate2", "G7":"gate5", "G8":"gate3", "G9":"gate1", "G10":"gate2", "G11":"gate4", "G12":"gate3", "H1":"gate2", "H2":"gate4", "H3":"gate5", "H4":"gate3", "H5":"gate1", "H6":"gate3", "H7":"gate2", "H8":"gate4", "H9":"", "H10":"", "H11":"", "H12":""}
two_pop_384w_dict = {"A1":"gate1","A2":"gate2","A3":"gate1","A4":"gate2","A5":"gate1","A6":"gate2","A7":"gate1","A8":"gate2","A9":"gate1","A10":"gate2","A11":"gate1","A12":"gate2","A13":"gate1","A14":"gate2","A15":"gate1","A16":"gate2","A17":"gate1","A18":"gate2","A19":"gate1","A20":"gate2","A21":"gate1","A22":"gate2","A23":"gate1","A24":"gate2","B1":"gate1","B2":"gate2","B3":"gate1","B4":"gate2","B5":"gate1","B6":"gate2","B7":"gate1","B8":"gate2","B9":"gate1","B10":"gate2","B11":"gate1","B12":"gate2","B13":"gate1","B14":"gate2","B15":"gate1","B16":"gate2","B17":"gate1","B18":"gate2","B19":"gate1","B20":"gate2","B21":"gate1","B22":"gate2","B23":"gate1","B24":"gate2","C1":"gate2","C2":"gate1","C3":"gate2","C4":"gate1","C5":"gate2","C6":"gate1","C7":"gate2","C8":"gate1","C9":"gate2","C10":"gate1","C11":"gate2","C12":"gate1","C13":"gate2","C14":"gate1","C15":"gate2","C16":"gate1","C17":"gate2","C18":"gate1","C19":"gate2","C20":"gate1","C21":"gate2","C22":"gate1","C23":"gate2","C24":"gate1","D1":"gate2","D2":"gate1","D3":"gate2","D4":"gate1","D5":"gate2","D6":"gate1","D7":"gate2","D8":"gate1","D9":"gate2","D10":"gate1","D11":"gate2","D12":"gate1","D13":"gate2","D14":"gate1","D15":"gate2","D16":"gate1","D17":"gate2","D18":"gate1","D19":"gate2","D20":"gate1","D21":"gate2","D22":"gate1","D23":"gate2","D24":"gate1","E1":"gate1","E2":"gate2","E3":"gate1","E4":"gate2","E5":"gate1","E6":"gate2","E7":"gate1","E8":"gate2","E9":"gate1","E10":"gate2","E11":"gate1","E12":"gate2","E13":"gate1","E14":"gate2","E15":"gate1","E16":"gate2","E17":"gate1","E18":"gate2","E19":"gate1","E20":"gate2","E21":"gate1","E22":"gate2","E23":"gate1","E24":"gate2","F1":"gate1","F2":"gate2","F3":"gate1","F4":"gate2","F5":"gate1","F6":"gate2","F7":"gate1","F8":"gate2","F9":"gate1","F10":"gate2","F11":"gate1","F12":"gate2","F13":"gate1","F14":"gate2","F15":"gate1","F16":"gate2","F17":"gate1","F18":"gate2","F19":"gate1","F20":"gate2","F21":"gate1","F22":"gate2","F23":"gate1","F24":"gate2","G1":"gate2","G2":"gate1","G3":"gate2","G4":"gate1","G5":"gate2","G6":"gate1","G7":"gate2","G8":"gate1","G9":"gate2","G10":"gate1","G11":"gate2","G12":"gate1","G13":"gate2","G14":"gate1","G15":"gate2","G16":"gate1","G17":"gate2","G18":"gate1","G19":"gate2","G20":"gate1","G21":"gate2","G22":"gate1","G23":"gate2","G24":"gate1","H1":"gate2","H2":"gate1","H3":"gate2","H4":"gate1","H5":"gate2","H6":"gate1","H7":"gate2","H8":"gate1","H9":"gate2","H10":"gate1","H11":"gate2","H12":"gate1","H13":"gate2","H14":"gate1","H15":"gate2","H16":"gate1","H17":"gate2","H18":"gate1","H19":"gate2","H20":"gate1","H21":"gate2","H22":"gate1","H23":"gate2","H24":"gate1","I1":"gate1","I2":"gate2","I3":"gate1","I4":"gate2","I5":"gate1","I6":"gate2","I7":"gate1","I8":"gate2","I9":"gate1","I10":"gate2","I11":"gate1","I12":"gate2","I13":"gate1","I14":"gate2","I15":"gate1","I16":"gate2","I17":"gate1","I18":"gate2","I19":"gate1","I20":"gate2","I21":"gate1","I22":"gate2","I23":"gate1","I24":"gate2","J1":"gate1","J2":"gate2","J3":"gate1","J4":"gate2","J5":"gate1","J6":"gate2","J7":"gate1","J8":"gate2","J9":"gate1","J10":"gate2","J11":"gate1","J12":"gate2","J13":"gate1","J14":"gate2","J15":"gate1","J16":"gate2","J17":"gate1","J18":"gate2","J19":"gate1","J20":"gate2","J21":"gate1","J22":"gate2","J23":"gate1","J24":"gate2","K1":"gate2","K2":"gate1","K3":"gate2","K4":"gate1","K5":"gate2","K6":"gate1","K7":"gate2","K8":"gate1","K9":"gate2","K10":"gate1","K11":"gate2","K12":"gate1","K13":"gate2","K14":"gate1","K15":"gate2","K16":"gate1","K17":"gate2","K18":"gate1","K19":"gate2","K20":"gate1","K21":"gate2","K22":"gate1","K23":"gate2","K24":"gate1","L1":"gate2","L2":"gate1","L3":"gate2","L4":"gate1","L5":"gate2","L6":"gate1","L7":"gate2","L8":"gate1","L9":"gate2","L10":"gate1","L11":"gate2","L12":"gate1","L13":"gate2","L14":"gate1","L15":"gate2","L16":"gate1","L17":"gate2","L18":"gate1","L19":"gate2","L20":"gate1","L21":"gate2","L22":"gate1","L23":"gate2","L24":"gate1","M1":"gate1","M2":"gate2","M3":"gate1","M4":"gate2","M5":"gate1","M6":"gate2","M7":"gate1","M8":"gate2","M9":"gate1","M10":"gate2","M11":"gate1","M12":"gate2","M13":"gate1","M14":"gate2","M15":"gate1","M16":"gate2","M17":"gate1","M18":"gate2","M19":"gate1","M20":"gate2","M21":"gate1","M22":"gate2","M23":"gate1","M24":"gate2","N1":"gate1","N2":"gate2","N3":"gate1","N4":"gate2","N5":"gate1","N6":"gate2","N7":"gate1","N8":"gate2","N9":"gate1","N10":"gate2","N11":"gate1","N12":"gate2","N13":"gate1","N14":"gate2","N15":"gate1","N16":"gate2","N17":"gate1","N18":"gate2","N19":"gate1","N20":"gate2","N21":"gate1","N22":"gate2","N23":"gate1","N24":"gate2","O1":"gate2","O2":"gate1","O3":"gate2","O4":"gate1","O5":"gate2","O6":"gate1","O7":"gate2","O8":"gate1","O9":"gate2","O10":"gate1","O11":"gate2","O12":"gate1","O13":"gate2","O14":"gate1","O15":"gate2","O16":"gate1","O17":"gate2","O18":"gate1","O19":"gate2","O20":"gate1","O21":"gate2","O22":"gate1","O23":"gate2","O24":"gate1","P1":"gate2","P2":"gate1","P3":"gate2","P4":"gate1","P5":"gate2","P6":"gate1","P7":"gate2","P8":"gate1","P9":"gate2","P10":"gate1","P11":"gate2","P12":"gate1","P13":"gate2","P14":"gate1","P15":"gate2","P16":"gate1","P17":"gate2","P18":"gate1","P19":"gate2","P20":"gate1","P21":"","P22":"","P23":"","P24":""}
three_pop_384w_dict = {"A1":"gate1","A2":"gate1","A3":"gate2","A4":"gate2","A5":"gate3","A6":"gate3","A7":"gate2","A8":"gate2","A9":"gate3","A10":"gate3","A11":"gate1","A12":"gate1","A13":"gate3","A14":"gate3","A15":"gate1","A16":"gate1","A17":"gate2","A18":"gate2","A19":"gate1","A20":"gate1","A21":"gate2","A22":"gate2","A23":"gate3","A24":"gate3","B1":"gate2","B2":"gate3","B3":"gate1","B4":"gate3","B5":"gate1","B6":"gate2","B7":"gate1","B8":"gate3","B9":"gate1","B10":"gate2","B11":"gate2","B12":"gate3","B13":"gate1","B14":"gate2","B15":"gate2","B16":"gate3","B17":"gate1","B18":"gate3","B19":"gate2","B20":"gate3","B21":"gate1","B22":"gate3","B23":"gate1","B24":"gate2","C1":"gate3","C2":"gate2","C3":"gate3","C4":"gate1","C5":"gate2","C6":"gate1","C7":"gate3","C8":"gate1","C9":"gate2","C10":"gate1","C11":"gate3","C12":"gate2","C13":"gate2","C14":"gate1","C15":"gate3","C16":"gate2","C17":"gate3","C18":"gate1","C19":"gate3","C20":"gate2","C21":"gate3","C22":"gate1","C23":"gate2","C24":"gate1","D1":"gate1","D2":"gate2","D3":"gate2","D4":"gate3","D5":"gate3","D6":"gate1","D7":"gate2","D8":"gate3","D9":"gate3","D10":"gate1","D11":"gate1","D12":"gate2","D13":"gate3","D14":"gate1","D15":"gate1","D16":"gate2","D17":"gate2","D18":"gate3","D19":"gate1","D20":"gate2","D21":"gate2","D22":"gate3","D23":"gate3","D24":"gate1","E1":"gate3","E2":"gate1","E3":"gate3","E4":"gate1","E5":"gate2","E6":"gate2","E7":"gate3","E8":"gate1","E9":"gate2","E10":"gate2","E11":"gate3","E12":"gate1","E13":"gate2","E14":"gate2","E15":"gate3","E16":"gate1","E17":"gate3","E18":"gate1","E19":"gate3","E20":"gate1","E21":"gate3","E22":"gate1","E23":"gate2","E24":"gate2","F1":"gate2","F2":"gate3","F3":"gate1","F4":"gate2","F5":"gate1","F6":"gate3","F7":"gate1","F8":"gate2","F9":"gate1","F10":"gate3","F11":"gate2","F12":"gate3","F13":"gate1","F14":"gate3","F15":"gate2","F16":"gate3","F17":"gate1","F18":"gate2","F19":"gate2","F20":"gate3","F21":"gate1","F22":"gate2","F23":"gate1","F24":"gate3","G1":"gate3","G2":"gate2","G3":"gate3","G4":"gate1","G5":"gate2","G6":"gate1","G7":"gate3","G8":"gate1","G9":"gate2","G10":"gate1","G11":"gate3","G12":"gate2","G13":"gate2","G14":"gate1","G15":"gate3","G16":"gate2","G17":"gate3","G18":"gate1","G19":"gate3","G20":"gate2","G21":"gate3","G22":"gate1","G23":"gate2","G24":"gate1","H1":"gate1","H2":"gate2","H3":"gate2","H4":"gate3","H5":"gate3","H6":"gate1","H7":"gate2","H8":"gate3","H9":"gate3","H10":"gate1","H11":"gate1","H12":"gate2","H13":"gate3","H14":"gate1","H15":"gate1","H16":"gate2","H17":"gate2","H18":"gate3","H19":"gate1","H20":"gate2","H21":"gate2","H22":"gate3","H23":"gate3","H24":"gate1","I1":"gate3","I2":"gate1","I3":"gate3","I4":"gate1","I5":"gate2","I6":"gate2","I7":"gate3","I8":"gate1","I9":"gate2","I10":"gate2","I11":"gate3","I12":"gate1","I13":"gate2","I14":"gate2","I15":"gate3","I16":"gate1","I17":"gate3","I18":"gate1","I19":"gate3","I20":"gate1","I21":"gate3","I22":"gate1","I23":"gate2","I24":"gate2","J1":"gate2","J2":"gate3","J3":"gate1","J4":"gate2","J5":"gate1","J6":"gate3","J7":"gate1","J8":"gate2","J9":"gate1","J10":"gate3","J11":"gate2","J12":"gate3","J13":"gate1","J14":"gate3","J15":"gate2","J16":"gate3","J17":"gate1","J18":"gate2","J19":"gate2","J20":"gate3","J21":"gate1","J22":"gate2","J23":"gate1","J24":"gate3","K1":"gate1","K2":"gate1","K3":"gate2","K4":"gate2","K5":"gate3","K6":"gate3","K7":"gate2","K8":"gate2","K9":"gate3","K10":"gate3","K11":"gate1","K12":"gate1","K13":"gate3","K14":"gate3","K15":"gate1","K16":"gate1","K17":"gate2","K18":"gate2","K19":"gate1","K20":"gate1","K21":"gate2","K22":"gate2","K23":"gate3","K24":"gate3","L1":"gate2","L2":"gate3","L3":"gate1","L4":"gate3","L5":"gate1","L6":"gate2","L7":"gate1","L8":"gate3","L9":"gate1","L10":"gate2","L11":"gate2","L12":"gate3","L13":"gate1","L14":"gate2","L15":"gate2","L16":"gate3","L17":"gate1","L18":"gate3","L19":"gate2","L20":"gate3","L21":"gate1","L22":"gate3","L23":"gate1","L24":"gate2","M1":"gate3","M2":"gate1","M3":"gate3","M4":"gate1","M5":"gate2","M6":"gate2","M7":"gate3","M8":"gate1","M9":"gate2","M10":"gate2","M11":"gate3","M12":"gate1","M13":"gate2","M14":"gate2","M15":"gate3","M16":"gate1","M17":"gate3","M18":"gate1","M19":"gate3","M20":"gate1","M21":"gate3","M22":"gate1","M23":"gate2","M24":"gate2","N1":"gate2","N2":"gate3","N3":"gate1","N4":"gate2","N5":"gate1","N6":"gate3","N7":"gate1","N8":"gate2","N9":"gate1","N10":"gate3","N11":"gate2","N12":"gate3","N13":"gate1","N14":"gate3","N15":"gate2","N16":"gate3","N17":"gate1","N18":"gate2","N19":"gate2","N20":"gate3","N21":"gate1","N22":"gate2","N23":"gate1","N24":"gate3","O1":"gate3","O2":"gate2","O3":"gate3","O4":"gate1","O5":"gate2","O6":"gate1","O7":"gate3","O8":"gate1","O9":"gate2","O10":"gate1","O11":"gate3","O12":"gate2","O13":"gate2","O14":"gate1","O15":"gate3","O16":"gate2","O17":"gate3","O18":"gate1","O19":"gate3","O20":"gate2","O21":"gate3","O22":"gate1","O23":"gate2","O24":"gate1","P1":"gate1","P2":"gate2","P3":"gate2","P4":"gate3","P5":"gate3","P6":"gate1","P7":"gate2","P8":"gate3","P9":"gate3","P10":"gate1","P11":"gate1","P12":"gate2","P13":"gate3","P14":"gate1","P15":"gate1","P16":"gate2","P17":"gate2","P18":"gate3","P19":"gate1","P20":"gate2","P21":"","P22":"","P23":"","P24":""}
four_pop_384w_dict = {"A1":"gate1","A2":"gate2","A3":"gate2","A4":"gate2","A5":"gate3","A6":"gate3","A7":"gate4","A8":"gate1","A9":"gate1","A10":"gate2","A11":"gate2","A12":"gate3","A13":"gate3","A14":"gate4","A15":"gate4","A16":"gate1","A17":"gate2","A18":"gate2","A19":"gate3","A20":"gate3","A21":"gate4","A22":"gate4","A23":"gate1","A24":"gate2","B1":"gate3","B2":"gate4","B3":"gate1","B4":"gate3","B5":"gate1","B6":"gate4","B7":"gate3","B8":"gate2","B9":"gate4","B10":"gate1","B11":"gate3","B12":"gate1","B13":"gate4","B14":"gate3","B15":"gate2","B16":"gate2","B17":"gate1","B18":"gate3","B19":"gate1","B20":"gate4","B21":"gate3","B22":"gate2","B23":"gate4","B24":"gate1","C1":"gate3","C2":"gate2","C3":"gate4","C4":"gate1","C5":"gate4","C6":"gate2","C7":"gate1","C8":"gate3","C9":"gate2","C10":"gate4","C11":"gate1","C12":"gate4","C13":"gate2","C14":"gate1","C15":"gate3","C16":"gate3","C17":"gate4","C18":"gate1","C19":"gate4","C20":"gate2","C21":"gate1","C22":"gate3","C23":"gate2","C24":"gate4","D1":"gate4","D2":"gate3","D3":"gate3","D4":"gate4","D5":"gate2","D6":"gate1","D7":"gate2","D8":"gate1","D9":"gate3","D10":"gate3","D11":"gate4","D12":"gate2","D13":"gate1","D14":"gate2","D15":"gate1","D16":"gate4","D17":"gate3","D18":"gate4","D19":"gate2","D20":"gate1","D21":"gate2","D22":"gate1","D23":"gate3","D24":"gate3","E1":"gate2","E2":"gate2","E3":"gate3","E4":"gate3","E5":"gate4","E6":"gate4","E7":"gate1","E8":"gate2","E9":"gate2","E10":"gate3","E11":"gate3","E12":"gate4","E13":"gate4","E14":"gate1","E15":"gate2","E16":"gate2","E17":"gate3","E18":"gate3","E19":"gate4","E20":"gate4","E21":"gate1","E22":"gate2","E23":"gate1","E24":"gate1","F1":"gate1","F2":"gate3","F3":"gate1","F4":"gate4","F5":"gate3","F6":"gate2","F7":"gate4","F8":"gate1","F9":"gate3","F10":"gate1","F11":"gate4","F12":"gate3","F13":"gate2","F14":"gate2","F15":"gate1","F16":"gate3","F17":"gate1","F18":"gate4","F19":"gate3","F20":"gate2","F21":"gate4","F22":"gate1","F23":"gate2","F24":"gate4","G1":"gate4","G2":"gate1","G3":"gate4","G4":"gate2","G5":"gate1","G6":"gate3","G7":"gate2","G8":"gate4","G9":"gate1","G10":"gate4","G11":"gate2","G12":"gate1","G13":"gate3","G14":"gate3","G15":"gate4","G16":"gate1","G17":"gate4","G18":"gate2","G19":"gate1","G20":"gate3","G21":"gate2","G22":"gate4","G23":"gate3","G24":"gate2","H1":"gate3","H2":"gate4","H3":"gate2","H4":"gate1","H5":"gate2","H6":"gate1","H7":"gate3","H8":"gate3","H9":"gate4","H10":"gate2","H11":"gate1","H12":"gate2","H13":"gate1","H14":"gate4","H15":"gate3","H16":"gate4","H17":"gate2","H18":"gate1","H19":"gate2","H20":"gate1","H21":"gate3","H22":"gate3","H23":"gate4","H24":"gate3","I1":"gate3","I2":"gate3","I3":"gate4","I4":"gate4","I5":"gate1","I6":"gate2","I7":"gate2","I8":"gate3","I9":"gate3","I10":"gate4","I11":"gate4","I12":"gate1","I13":"gate2","I14":"gate2","I15":"gate3","I16":"gate3","I17":"gate4","I18":"gate4","I19":"gate1","I20":"gate2","I21":"gate1","I22":"gate1","I23":"gate2","I24":"gate2","J1":"gate1","J2":"gate4","J3":"gate3","J4":"gate2","J5":"gate4","J6":"gate1","J7":"gate3","J8":"gate1","J9":"gate4","J10":"gate3","J11":"gate2","J12":"gate2","J13":"gate1","J14":"gate3","J15":"gate1","J16":"gate4","J17":"gate3","J18":"gate2","J19":"gate4","J20":"gate1","J21":"gate2","J22":"gate4","J23":"gate1","J24":"gate3","K1":"gate4","K2":"gate2","K3":"gate1","K4":"gate3","K5":"gate2","K6":"gate4","K7":"gate1","K8":"gate4","K9":"gate2","K10":"gate1","K11":"gate3","K12":"gate3","K13":"gate4","K14":"gate1","K15":"gate4","K16":"gate2","K17":"gate1","K18":"gate3","K19":"gate2","K20":"gate4","K21":"gate3","K22":"gate2","K23":"gate4","K24":"gate1","L1":"gate2","L2":"gate1","L3":"gate2","L4":"gate1","L5":"gate3","L6":"gate3","L7":"gate4","L8":"gate2","L9":"gate1","L10":"gate2","L11":"gate1","L12":"gate4","L13":"gate3","L14":"gate4","L15":"gate2","L16":"gate1","L17":"gate2","L18":"gate1","L19":"gate3","L20":"gate3","L21":"gate4","L22":"gate3","L23":"gate3","L24":"gate4","M1":"gate4","M2":"gate4","M3":"gate1","M4":"gate2","M5":"gate2","M6":"gate3","M7":"gate3","M8":"gate4","M9":"gate3","M10":"gate3","M11":"gate4","M12":"gate4","M13":"gate1","M14":"gate2","M15":"gate2","M16":"gate3","M17":"gate3","M18":"gate4","M19":"gate3","M20":"gate3","M21":"gate4","M22":"gate4","M23":"gate1","M24":"gate2","N1":"gate3","N2":"gate2","N3":"gate4","N4":"gate1","N5":"gate3","N6":"gate1","N7":"gate4","N8":"gate3","N9":"gate1","N10":"gate4","N11":"gate3","N12":"gate2","N13":"gate4","N14":"gate1","N15":"gate3","N16":"gate1","N17":"gate4","N18":"gate3","N19":"gate1","N20":"gate4","N21":"gate3","N22":"gate2","N23":"gate4","N24":"gate1","O1":"gate1","O2":"gate3","O3":"gate2","O4":"gate4","O5":"gate1","O6":"gate4","O7":"gate2","O8":"gate1","O9":"gate4","O10":"gate2","O11":"gate1","O12":"gate3","O13":"gate2","O14":"gate4","O15":"gate1","O16":"gate4","O17":"gate2","O18":"gate1","O19":"gate4","O20":"gate2","O21":"gate1","O22":"gate3","O23":"gate2","O24":"gate4","P1":"gate2","P2":"gate1","P3":"gate3","P4":"gate3","P5":"gate4","P6":"gate2","P7":"gate1","P8":"gate2","P9":"gate2","P10":"gate1","P11":"gate2","P12":"gate1","P13":"gate3","P14":"gate3","P15":"gate4","P16":"gate2","P17":"gate1","P18":"gate2","P19":"gate2","P20":"gate1","P21":"","P22":"","P23":"","P24":""}

#aList is going to contain the list of gates generated by the user
aList=[] 
#counts the number of gates (gatecounter.pop) and saves them (aList)
def gatecounter():
    SetupFile = open(max(glob.iglob('*.xml'), key=os.path.getctime))
    for line in SetupFile:
        line = line.rstrip()
        x = re.findall('<Gate>\S+', line)
        y = ['<Gate></Gate>']
        try:
            SampleName = re.search('<Gate>(.+?)</Gate>', line).group(1)
        except AttributeError:
            found = ''
        if len(x) > 0 and y != x :
            aList.append(SampleName)
    gatecounter.pop = len(aList)
    SetupFile.close()

#reorders the dictionary to match the well order in the sort layout file
def gatemultiplier(replacer):
    SetupFile = open(max(glob.iglob('*.xml'), key=os.path.getctime))
    gatemultiplier.neworder =[]
    for dictline in SetupFile:
        for dictkey in replacer.keys():
            if "<Name>"+dictkey+"</Name>" in dictline:
                gatemultiplier.neworder.append(replacer[dictkey])
    SetupFile.close()
#replaces the dictionary gate names with the user gate names and writes the file
def filemaker(replacements):
    SetupFile = fileinput.input(files=(max(glob.iglob('*.xml'), key=os.path.getctime)), inplace=1)
    text = "<Gate>"
    gate01="\t\t<Gate>"
    gate02="</Gate>\n"
    ii=0
    for line in SetupFile:
        if text in line:
            line = gate01+gatemultiplier.neworder[ii]+gate02
            ii+=1
            for i in replacements.keys():
                line = line.replace(i, replacements[i])
        sys.stdout.write (line,)
    SetupFile.close()
#combines all of the above, but in one function plus at the end creates a new name for the sort layout
def multi_file(replacer, copyname, replacements, newname):
    SetupFile = open(copyname)
    gatemultiplier.neworder =[]
    for dictline in SetupFile:
        for dictkey in replacer.keys():
            if "<Name>"+dictkey+"</Name>" in dictline:
                gatemultiplier.neworder.append(replacer[dictkey])
    SetupFile.close()
    SetupFile = fileinput.input(copyname, inplace=1)
    text = "<Gate>"
    gate01="\t\t<Gate>"
    gate02="</Gate>\n"
    ii=0
    for line in SetupFile:
        if text in line:
            line = gate01+gatemultiplier.neworder[ii]+gate02
            ii+=1
            for i in replacements.keys():
                line = line.replace(i, replacements[i])
        sys.stdout.write (line,)
    SetupFile.close()
    file=open(copyname, "r")
    for line in file:
        match=re.search('<Name>(.+)</Name>',line)
        if match:
            oldname = (match.group(1))
            break
    file.close()
    for line in fileinput.FileInput(copyname,inplace=1):
        line = line.replace(oldname,oldname+":"+newname)
        sys.stdout.write (line,)
#used in the custom layouts, simal to multi_file
def singlefile(replacer, gatenum):
    SetupFile = open(max(glob.iglob('*.xml'), key=os.path.getctime))
    gatemultiplier.neworder =[]
    for dictline in SetupFile:
        for dictkey in replacer.keys():
            if "<Name>"+dictkey+"</Name>" in dictline:
                gatemultiplier.neworder.append(replacer[dictkey])
    SetupFile.close()
    if gatenum ==2:
        replacements={'gate1':str(aList[0]), 'gate2':str(aList[1])}
    elif gatenum ==3:
        replacements={'gate1':str(aList[0]), 'gate2':str(aList[1]), 'gate3':str(aList[2])}
    elif gatenum ==4:
        replacements={'gate1':str(aList[0]), 'gate2':str(aList[1]), 'gate3':str(aList[2]), 'gate4':str(aList[3])}
    elif gatenum ==5:
        replacements={'gate1':str(aList[0]), 'gate2':str(aList[1]), 'gate3':str(aList[2]), 'gate4':str(aList[3]), 'gate5':str(aList[4])}
    else:
        print("Incompatible plate formatting.  Number of gates=", gatecounter.pop)
        time.sleep(5)
    SetupFile = fileinput.input(files=(max(glob.iglob('*.xml'), key=os.path.getctime)), inplace=1)
    text = "<Gate>"
    #new_text = "\t\t<Gate></Gate>\n"
    gate1="\t\t<Gate>"
    gate2="</Gate>\n"
    ii=0
    for line in SetupFile:
        if text in line:
            line = gate1+gatemultiplier.neworder[ii]+gate2
            ii+=1
            for i in replacements.keys():
                line = line.replace(i, replacements[i])
        sys.stdout.write (line,)
    SetupFile.close()
def custom_multifile(gate_quant, cust_template):
    if gate_quant == 1:
        print ("No need for latin square. Number of gates=", gate_quant)
        time.sleep(5)
    elif gate_quant == 2:
        for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
        copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
        copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
        miss0={'gate1':str(aList[0]), 'gate2':""}
        miss1={'gate1':"", 'gate2':str(aList[1])}
        multi_file(cust_template, copy1, miss0, (aList[0]))
        multi_file(cust_template, copy2, miss1, (aList[1]))
    elif gate_quant == 3:
        for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
        copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
        copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
        copy3=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_3.xml")
        miss0={'gate1':str(aList[0]), 'gate2':"", 'gate3':""}
        miss1={'gate1':"", 'gate2':str(aList[1]), 'gate3':""}
        miss2={'gate1':"", 'gate2':"", 'gate3':str(aList[2])}
        multi_file(cust_template, copy1, miss0, (aList[0]))
        multi_file(cust_template, copy2, miss1, (aList[1]))
        multi_file(cust_template, copy3, miss2, (aList[2]))
    elif gate_quant == 4:
        for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
        copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
        copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
        copy3=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_3.xml")
        copy4=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_4.xml")
        miss0={'gate1':str(aList[0]), 'gate2':"", 'gate3':"", 'gate4':""}
        miss1={'gate1':"", 'gate2':str(aList[1]), 'gate3':"", 'gate4':""}
        miss2={'gate1':"", 'gate2':"", 'gate3':str(aList[2]), 'gate4':""}
        miss3={'gate1':"", 'gate2':"", 'gate3':"", 'gate4':str(aList[3])}
        multi_file(cust_template, copy1, miss0, (aList[0]))
        multi_file(cust_template, copy2, miss1, (aList[1]))
        multi_file(cust_template, copy3, miss2, (aList[2]))
        multi_file(cust_template, copy4, miss3, (aList[3]))    
    elif gate_quant == 5:
        for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
        copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
        copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
        copy3=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_3.xml")
        copy4=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_4.xml")
        copy5=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_5.xml")
        miss0={'gate1':str(aList[0]), 'gate2':"", 'gate3':"", 'gate4':"", 'gate5':""}
        miss1={'gate1':"", 'gate2':str(aList[1]), 'gate3':"", 'gate4':"", 'gate5':""}
        miss2={'gate1':"", 'gate2':"", 'gate3':str(aList[2]), 'gate4':"", 'gate5':""}
        miss3={'gate1':"", 'gate2':"", 'gate3':"", 'gate4':str(aList[3]), 'gate5':""}
        miss4={'gate1':"", 'gate2':"", 'gate3':"", 'gate4':"", 'gate5':str(aList[4])}
        multi_file(cust_template, copy1, miss0, (aList[0]))
        multi_file(cust_template, copy2, miss1, (aList[1]))
        multi_file(cust_template, copy3, miss2, (aList[2]))
        multi_file(cust_template, copy4, miss3, (aList[3]))
        multi_file(cust_template, copy5, miss4, (aList[4]))    
    else:
        print("Incompatible plate formatting. Number of gates=", gate_quant)
        time.sleep(5)
#warn the user about the bug! and ask whether they want one output or multiple outputs
print ("Latin Square Generator for BD Influx running Sortware v1.2.0.142\nVersion 1.81 APR 2018\n\nAuthor:\nChristopher Hall (Wellcome Sanger Institute)\nLatin Square designs by Stephan Lorenz (Wellcome Sanger Instiute)\n\nThis program will use the latest sort layout in the folder\n")
question = str(input("Type '1' for single sample with many gates\nType '2' for many samples and a single gate\nType '3' for custom layouts\n"))
#one file
if question == "1":
#counts the number of lines in the xml file to guess if it is 96 or 384 well template    
    Startup = open(max(glob.iglob('*.xml'), key=os.path.getctime))
    lines = sum(1 for line in Startup)
    Startup.close()
    if lines < 600:
#96 well program
        gatecounter()
        if gatecounter.pop ==1:
            print ("No need for latin square. Number of gates=", gatecounter.pop)
            time.sleep(5)
        elif gatecounter.pop ==2:
            gatemultiplier(two_pop_96w_dict)
            rep_2={'gate1':str(aList[0]), 'gate2':str(aList[1])}
            filemaker(rep_2)
        elif gatecounter.pop ==3:        
            gatemultiplier(three_pop_96w_dict)
            rep_3={'gate1':str(aList[0]), 'gate2':str(aList[1]), 'gate3':str(aList[2])}
            filemaker(rep_3)
        elif gatecounter.pop ==4:
            gatemultiplier(four_pop_96w_dict)
            rep_4={'gate1':str(aList[0]), 'gate2':str(aList[1]), 'gate3':str(aList[2]), 'gate4':str(aList[3])}
            filemaker(rep_4)
        elif gatecounter.pop ==5:
            gatemultiplier(five_pop_96w_dict)
            rep_5={'gate1':str(aList[0]), 'gate2':str(aList[1]), 'gate3':str(aList[2]), 'gate4':str(aList[3]), 'gate5':str(aList[4])}
            filemaker(rep_5)
        else:
            print("Incompatible plate formatting.  Number of gates=", gatecounter.pop)
            time.sleep(5)
    elif lines >600:
#384 well program
        gatecounter()
        if gatecounter.pop ==1:
            print ("No need for latin square. Number of gates=", gatecounter.pop)
            time.sleep(5)
        elif gatecounter.pop ==2:
            gatemultiplier(two_pop_384w_dict)
            rep_2={'gate1':str(aList[0]), 'gate2':str(aList[1])}
            filemaker(rep_2)
        elif gatecounter.pop ==3:
            gatemultiplier(three_pop_384w_dict)
            rep_3={'gate1':str(aList[0]), 'gate2':str(aList[1]), 'gate3':str(aList[2])}
            filemaker(rep_3)
        elif gatecounter.pop ==4:
            gatemultiplier(four_pop_384w_dict)
            rep_4={'gate1':str(aList[0]), 'gate2':str(aList[1]), 'gate3':str(aList[2]), 'gate4':str(aList[3])}
            filemaker(rep_4)
        else:
            print("Incompatible plate formatting. Number of gates=", gatecounter.pop)
            time.sleep(5)
    else:
        print("Incompatible plate formatting")
        time.sleep(5)
#multiple files
elif question == "2":
    Startup = open(max(glob.iglob('*.xml'), key=os.path.getctime))
    lines = sum(1 for line in Startup)
    Startup.close()
    if lines < 600:
        gatecounter()
        if gatecounter.pop ==1:
            print ("No need for latin square. Number of gates=", gatecounter.pop)
            time.sleep(5)
        elif gatecounter.pop ==2:
            #deleting the previous files, makes life easier
            for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
            #creates new files
            copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
            copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
            miss0={'gate1':str(aList[0]), 'gate2':""}
            miss1={'gate1':"", 'gate2':str(aList[1])}
            multi_file(two_pop_96w_dict, copy1, miss0, (aList[0]))
            multi_file(two_pop_96w_dict, copy2, miss1, (aList[1]))
        elif gatecounter.pop ==3:        
            for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
            copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
            copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
            copy3=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_3.xml")
            miss0={'gate1':str(aList[0]), 'gate2':"", 'gate3':""}
            miss1={'gate1':"", 'gate2':str(aList[1]), 'gate3':""}
            miss2={'gate1':"", 'gate2':"", 'gate3':str(aList[2])}
            multi_file(three_pop_96w_dict, copy1, miss0, (aList[0]))
            multi_file(three_pop_96w_dict, copy2, miss1, (aList[1]))
            multi_file(three_pop_96w_dict, copy3, miss2, (aList[2]))
        elif gatecounter.pop ==4:
            for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
            copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
            copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
            copy3=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_3.xml")
            copy4=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_4.xml")
            miss0={'gate1':str(aList[0]), 'gate2':"", 'gate3':"", 'gate4':""}
            miss1={'gate1':"", 'gate2':str(aList[1]), 'gate3':"", 'gate4':""}
            miss2={'gate1':"", 'gate2':"", 'gate3':str(aList[2]), 'gate4':""}
            miss3={'gate1':"", 'gate2':"", 'gate3':"", 'gate4':str(aList[3])}
            multi_file(four_pop_96w_dict, copy1, miss0, (aList[0]))
            multi_file(four_pop_96w_dict, copy2, miss1, (aList[1]))
            multi_file(four_pop_96w_dict, copy3, miss2, (aList[2]))
            multi_file(four_pop_96w_dict, copy4, miss3, (aList[3]))
        elif gatecounter.pop ==5:
            for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
            copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
            copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
            copy3=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_3.xml")
            copy4=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_4.xml")
            copy5=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_5.xml")
            miss0={'gate1':str(aList[0]), 'gate2':"", 'gate3':"", 'gate4':"", 'gate5':""}
            miss1={'gate1':"", 'gate2':str(aList[1]), 'gate3':"", 'gate4':"", 'gate5':""}
            miss2={'gate1':"", 'gate2':"", 'gate3':str(aList[2]), 'gate4':"", 'gate5':""}
            miss3={'gate1':"", 'gate2':"", 'gate3':"", 'gate4':str(aList[3]), 'gate5':""}
            miss4={'gate1':"", 'gate2':"", 'gate3':"", 'gate4':"", 'gate5':str(aList[4])}
            multi_file(five_pop_96w_dict, copy1, miss0, (aList[0]))
            multi_file(five_pop_96w_dict, copy2, miss1, (aList[1]))
            multi_file(five_pop_96w_dict, copy3, miss2, (aList[2]))
            multi_file(five_pop_96w_dict, copy4, miss3, (aList[3]))
            multi_file(five_pop_96w_dict, copy5, miss4, (aList[4]))
        else:
            print("Incompatible plate formatting.  Number of gates=", gatecounter.pop)
            time.sleep(5)
    elif lines >600:
        gatecounter()
        if gatecounter.pop ==1:
            print ("No need for latin square. Number of gates=", gatecounter.pop)
            time.sleep(5)
        elif gatecounter.pop ==2:
            for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
            copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
            copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
            gatecounter()
            miss0={'gate1':str(aList[0]), 'gate2':""}
            miss1={'gate1':"", 'gate2':str(aList[1])}
            multi_file(two_pop_384w_dict, copy1, miss0, (aList[0]))
            multi_file(two_pop_384w_dict, copy2, miss1, (aList[1]))
        elif gatecounter.pop ==3:
            for f in glob.glob ('LatSqMultiSample/*.xml'):
                    os.remove (f)
            copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
            copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
            copy3=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_3.xml")
            gatecounter()
            miss0={'gate1':str(aList[0]), 'gate2':"", 'gate3':""}
            miss1={'gate1':"", 'gate2':str(aList[1]), 'gate3':""}
            miss2={'gate1':"", 'gate2':"", 'gate3':str(aList[2])}
            multi_file(three_pop_384w_dict, copy1, miss0, (aList[0]))
            multi_file(three_pop_384w_dict, copy2, miss1, (aList[1]))
            multi_file(three_pop_384w_dict, copy3, miss2, (aList[2]))
            
        elif gatecounter.pop ==4:
            for f in glob.glob ('LatSqMultiSample/*.xml'):
                os.remove (f)
            copy1=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_1.xml")
            copy2=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_2.xml")
            copy3=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_3.xml")
            copy4=shutil.copy((max(glob.iglob('*.xml'), key=os.path.getctime)), "LatSqMultiSample/new_sort_4.xml")
            gatecounter()
            miss0={'gate1':str(aList[0]), 'gate2':"", 'gate3':"", 'gate4':""}
            miss1={'gate1':"", 'gate2':str(aList[1]), 'gate3':"", 'gate4':""}
            miss2={'gate1':"", 'gate2':"", 'gate3':str(aList[2]), 'gate4':""}
            miss3={'gate1':"", 'gate2':"", 'gate3':"", 'gate4':str(aList[3])}
            multi_file(four_pop_384w_dict, copy1, miss0, (aList[0]))
            multi_file(four_pop_384w_dict, copy2, miss1, (aList[1]))
            multi_file(four_pop_384w_dict, copy3, miss2, (aList[2]))
            multi_file(four_pop_384w_dict, copy4, miss3, (aList[3]))
    else:
        print("Incompatible plate formatting. Number of gates=", gatecounter.pop)
        time.sleep(5)
#custom templates option.  Use the xlsx document as template
elif question =="3":
    import openpyxl as op
    from tkinter import filedialog as fd
    from tkinter import Tk
    root = Tk()
    root.title("Open Excel file")
    spreadsheet = fd.askopenfilename()
    root.destroy()
    layout=[]
    keys_96=['A1','A2','A3','A4','A5','A6','A7','A8','A9','A10','A11','A12','B1','B2','B3','B4','B5','B6','B7','B8','B9','B10','B11','B12','C1','C2','C3','C4','C5','C6','C7','C8','C9','C10','C11','C12','D1','D2','D3','D4','D5','D6','D7','D8','D9','D10','D11','D12','E1','E2','E3','E4','E5','E6','E7','E8','E9','E10','E11','E12','F1','F2','F3','F4','F5','F6','F7','F8','F9','F10','F11','F12','G1','G2','G3','G4','G5','G6','G7','G8','G9','G10','G11','G12','H1','H2','H3','H4','H5','H6','H7','H8','H9','H10','H11','H12']
    keys_384=['A1','A2','A3','A4','A5','A6','A7','A8','A9','A10','A11','A12','A13','A14','A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','B1','B2','B3','B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','C1','C2','C3','C4','C5','C6','C7','C8','C9','C10','C11','C12','C13','C14','C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','D1','D2','D3','D4','D5','D6','D7','D8','D9','D10','D11','D12','D13','D14','D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','E1','E2','E3','E4','E5','E6','E7','E8','E9','E10','E11','E12','E13','E14','E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','F1','F2','F3','F4','F5','F6','F7','F8','F9','F10','F11','F12','F13','F14','F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','G1','G2','G3','G4','G5','G6','G7','G8','G9','G10','G11','G12','G13','G14','G15','G16','G17','G18','G19','G20','G21','G22','G23','G24','H1','H2','H3','H4','H5','H6','H7','H8','H9','H10','H11','H12','H13','H14','H15','H16','H17','H18','H19','H20','H21','H22','H23','H24','I1','I2','I3','I4','I5','I6','I7','I8','I9','I10','I11','I12','I13','I14','I15','I16','I17','I18','I19','I20','I21','I22','I23','I24','J1','J2','J3','J4','J5','J6','J7','J8','J9','J10','J11','J12','J13','J14','J15','J16','J17','J18','J19','J20','J21','J22','J23','J24','K1','K2','K3','K4','K5','K6','K7','K8','K9','K10','K11','K12','K13','K14','K15','K16','K17','K18','K19','K20','K21','K22','K23','K24','L1','L2','L3','L4','L5','L6','L7','L8','L9','L10','L11','L12','L13','L14','L15','L16','L17','L18','L19','L20','L21','L22','L23','L24','M1','M2','M3','M4','M5','M6','M7','M8','M9','M10','M11','M12','M13','M14','M15','M16','M17','M18','M19','M20','M21','M22','M23','M24','N1','N2','N3','N4','N5','N6','N7','N8','N9','N10','N11','N12','N13','N14','N15','N16','N17','N18','N19','N20','N21','N22','N23','N24','O1','O2','O3','O4','O5','O6','O7','O8','O9','O10','O11','O12','O13','O14','O15','O16','O17','O18','O19','O20','O21','O22','O23','O24','P1','P2','P3','P4','P5','P6','P7','P8','P9','P10','P11','P12','P13','P14','P15','P16','P17','P18','P19','P20','P21','P22','P23','P24']
    #make a custom dictionary of well names and gate names.  I want to 1/2 this section but I cannot work out how to pass the coordinates to openpyxl :(
    cust_question = str(input("Copy the Excel document to the layout folder\n\nType '1' for 96 well layout \nType '2' for 384 well layout\n"))
    wb = op.load_workbook(filename = spreadsheet)
    if cust_question == '1':
        ws = wb['96']
        plate=ws['B2':'M9']
        for row in plate:
            for cell in row:
                if cell.value is not None:
                    layout.append(cell.value)
                else:
                    layout.append('')
        customdict=dict(zip(keys_96, layout))
    elif cust_question == '2':
        ws = wb['384']
        plate=ws['B2':'Y17']
        for row in plate:
            for cell in row:
                if cell.value is not None:
                    layout.append(cell.value)
                else:
                    layout.append('')
        customdict=dict(zip(keys_384, layout))
    else:
        print('unrecognised input.  Please enter 1 or 2')
    #run the script to make the files
    cust_oneormore = str(input("Type '1' for single sample with many gates\nType '2' for many samples and a single gate\n"))
    gatecounter()
    if cust_oneormore == "1":
        singlefile(customdict, gatecounter.pop)
    elif cust_oneormore =="2":
        custom_multifile(gatecounter.pop, customdict)

else:
    print ("Not understood, please restart and try again")
    time.sleep(5)
print("Finished and closing")
time.sleep(1)
